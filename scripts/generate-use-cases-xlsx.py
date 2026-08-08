"""Generate PROCHECK-User-Use-Cases.xlsx from backend-audit.json + client docs."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("f:/Munas/procheeck/project_doc")
AUDIT = ROOT / "backend-audit.json"
OUT = ROOT / "PROCHECK-User-Use-Cases.xlsx"

audit = json.loads(AUDIT.read_text(encoding="utf-8"))

# ---------- styling ----------
NAVY = "0F1725"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STATUS_FILL = {
    "DONE": PatternFill("solid", fgColor="DCFCE7"),
    "PARTIAL": PatternFill("solid", fgColor="FEF3C7"),
    "TODO": PatternFill("solid", fgColor="DBEAFE"),
    "BLOCKED": PatternFill("solid", fgColor="FEE2E2"),
    "CLARIFY": PatternFill("solid", fgColor="EDE9FE"),
}
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_status(ws, col_idx, start_row=2):
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
        for c in row:
            v = (c.value or "").strip() if isinstance(c.value, str) else ""
            if v in STATUS_FILL:
                c.fill = STATUS_FILL[v]
                c.font = Font(bold=True)


def wrap_all(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER


wb = Workbook()

# ============= SHEET 1: Overview =============
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False

ws["A1"] = "PROCHECK Safety - User Use Cases & Backend Scope"
ws["A1"].font = Font(bold=True, size=20, color=NAVY)
ws.merge_cells("A1:D1")

rows = [
    (),
    ("Client:", "Alejandra Ibarra Nunez (Ale)"),
    ("Client email:", "ale@procheeck.com"),
    ("Project:", "PROCHECK Safety - Mexican occupational safety training and certification platform"),
    ("Prepared by:", "Munas"),
    ("Date:", "2026-07-19"),
    ("Contact:", "Munas"),
    (),
    ("Purpose",),
    ("Lock backend scope by mapping every user, use case, and gap. Any change from here on must be quoted separately.",),
    (),
    ("Summary",),
    ("Roles:", "8 (Principal Admin, Vendedor, Capacitador, Client, Client Admin, Subcontractor, Employee, Public Verifier)"),
    ("Use cases:", "68"),
    ("Endpoints:", "77"),
    ("Entities:", "19"),
    ("Modules:", "16"),
    (),
    ("Top 5 CRITICAL gaps",),
    ("1", "Vendedor role not implemented (schema + endpoints + guards + UI)"),
    ("2", "Capacitador (instructor) role not implemented (schema + endpoints + guards + UI)"),
    ("3", "No CFDI/SAT invoicing endpoint - blocks Mexican tax compliance"),
    ("4", "No payment provider webhook - payments stuck in pending"),
    ("5", "No refresh token / session revocation - security risk for a regulated LMS"),
    (),
    ("Status legend",),
    ("DONE", "Implemented and verified in code"),
    ("PARTIAL", "Partial implementation, gaps remain"),
    ("TODO", "Not started, planned in scope"),
    ("BLOCKED", "Waiting on client decision or external dependency"),
    ("CLARIFY", "Requirement still ambiguous"),
]
r = 2
for row in rows:
    if not row:
        r += 1
        continue
    for i, val in enumerate(row):
        c = ws.cell(row=r, column=i + 1, value=val)
        if i == 0:
            c.font = Font(bold=True)
    # section headings
    if len(row) == 1:
        ws.cell(row=r, column=1).font = Font(bold=True, size=13, color=NAVY)
    # status legend colouring
    if row and row[0] in STATUS_FILL:
        ws.cell(row=r, column=1).fill = STATUS_FILL[row[0]]
        ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1

set_widths(ws, [22, 90, 20, 20])

# ============= SHEET 2: Roles =============
ws = wb.create_sheet("Roles")
headers = ["Role", "Codename", "Category", "Description", "Login", "Isolation Rule", "Backend Status"]
ws.append(headers)

roles_rows = [
    ("Principal Admin", "PRINCIPAL_ADMIN", "Internal", "PROCHECK super admin. Full platform access across all tenants.", "Yes", "No isolation - can see all companies, users, courses, certificates", "DONE"),
    ("Vendedor (NEW)", "VENDEDOR", "Internal", "Sales rep. Creates leads, closes deals, tracks commissions.", "Yes", "Sees only own leads and commissions; can view assigned client companies", "TODO - NEW role added via client annotations. Not yet in schema or code."),
    ("Capacitador (NEW)", "CAPACITADOR", "Internal", "Instructor. Delivers courses, records attendance, signs off completions.", "Yes", "Sees only assigned cohorts, courses, and enrolled students", "TODO - NEW role added via client annotations. Not yet in schema or code."),
    ("Client", "CLIENT", "External", "Company-level owner. Buys seats, oversees the whole client account.", "Yes", "Sees only own company data (users, enrollments, certs, invoices)", "DONE"),
    ("Client Admin", "CLIENT_ADMIN", "External", "Delegated administrator inside a client company. Manages employees and subcontractors.", "Yes", "Sees only own company + linked subcontractor companies", "DONE"),
    ("Subcontractor", "SUBCONTRACTOR", "External", "Third-party company linked to a client. Their employees can be enrolled by the client.", "Yes", "Sees only own company data; parent client can enroll them", "DONE"),
    ("Employee", "EMPLOYEE", "External", "End learner. Takes courses, quizzes, receives certificates.", "Yes", "Sees only own enrollments, quiz results, and certificates", "DONE"),
    ("Public Verifier", "PUBLIC", "Public", "Anyone verifying a DC-3 certificate via folio/QR. No account required.", "No", "Read-only public verification endpoint, no PII exposure", "PARTIAL - endpoint exists, QR verification flow not documented"),
]
for r in roles_rows:
    ws.append(r)

style_header(ws, len(headers))
set_widths(ws, [22, 20, 14, 55, 10, 50, 45])
wrap_all(ws)
apply_status(ws, 7)

# ============= SHEET 3: Use Cases =============
ws = wb.create_sheet("Use Cases")
headers = ["ID", "Role", "Use Case", "Trigger", "Steps", "Priority", "Backend Status", "Endpoint(s)", "Gap Notes"]
ws.append(headers)

PRIO_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "MED": 2, "LOW": 3}
use_cases = sorted(
    audit["useCases"],
    key=lambda u: (u.get("role", ""), PRIO_ORDER.get((u.get("priority") or "").upper(), 9)),
)
for uc in use_cases:
    ws.append([
        uc.get("id"),
        uc.get("role", ""),
        uc.get("useCase", ""),
        uc.get("trigger", ""),
        "; ".join(uc.get("steps", []) or []),
        uc.get("priority", ""),
        uc.get("backendStatus", ""),
        "; ".join(uc.get("endpoints", []) or []),
        uc.get("gaps", "") if isinstance(uc.get("gaps"), str) else "; ".join(uc.get("gaps", []) or []),
    ])

style_header(ws, len(headers))
set_widths(ws, [6, 18, 40, 32, 60, 10, 14, 45, 45])
wrap_all(ws)
apply_status(ws, 7)

# ============= SHEET 4: Endpoints =============
ws = wb.create_sheet("Endpoints")
headers = ["Module", "Method", "Path", "Purpose", "Auth Requirement", "Request Body", "Response Keys", "Status", "Notes"]
ws.append(headers)

endpoints = []
for m in audit["modules"]:
    for ep in m.get("endpoints", []):
        endpoints.append((m["name"], ep))
endpoints.sort(key=lambda x: (x[0], x[1].get("path", "")))

for mod_name, ep in endpoints:
    req = ep.get("requestBody", [])
    if isinstance(req, list):
        req_str = ", ".join(req) if req else ""
    else:
        req_str = str(req)
    resp = ep.get("responseKeys", [])
    resp_str = ", ".join(resp) if isinstance(resp, list) else str(resp)
    ws.append([
        mod_name,
        ep.get("method", ""),
        ep.get("path", ""),
        ep.get("purpose", ""),
        ep.get("auth", ""),
        req_str,
        resp_str,
        ep.get("status", ""),
        ep.get("notes", ""),
    ])

style_header(ws, len(headers))
set_widths(ws, [16, 10, 42, 40, 20, 40, 40, 12, 35])
wrap_all(ws)
apply_status(ws, 8)

# ============= SHEET 5: Data Model =============
ws = wb.create_sheet("Data Model")
headers = ["Entity", "Table", "Key Fields", "Relationships", "Read Access", "Write Access", "Notes"]
ws.append(headers)

# heuristic write access: Principal Admin always; specific others per entity name
WRITE_HINTS = {
    "Role": "Principal Admin (seed only)",
    "User": "Principal Admin; Client Admin (own company); Client (own company)",
    "Company": "Principal Admin; Client (own record)",
    "Course": "Principal Admin",
    "Enrollment": "Principal Admin; Client Admin; Client",
    "Quiz": "Principal Admin",
    "QuizAttempt": "Employee (self); auto-graded",
    "Certificate": "Principal Admin; system (auto-issue)",
    "Invoice": "Principal Admin; system (CFDI stamp - MISSING)",
    "Payment": "Principal Admin; system (webhook - MISSING)",
    "Upload": "All authenticated (own scope)",
    "LibraryDoc": "Principal Admin",
    "Notification": "system",
    "SupportTicket": "All authenticated (own tickets)",
    "AuditLog": "system (write-only)",
    "Mail": "system",
}
for ent in audit["entities"]:
    users = ent.get("usedByRoles", []) or []
    read_access = ", ".join(users) if users else "Principal Admin"
    write = WRITE_HINTS.get(ent["name"], "Principal Admin")
    key_fields = ", ".join(
        f.get("name", "") for f in ent.get("fields", [])[:6]
    )
    rels = "; ".join(ent.get("relationships", []) or [])
    ws.append([
        ent["name"],
        ent.get("table", ""),
        key_fields,
        rels,
        read_access,
        write,
        "",
    ])

style_header(ws, len(headers))
set_widths(ws, [18, 20, 40, 45, 45, 45, 30])
wrap_all(ws)

# ============= SHEET 6: Gaps and Blockers =============
ws = wb.create_sheet("Gaps and Blockers")

ws["A1"] = "A. Critical gaps (must fix before launch)"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = HEADER_FILL
ws.merge_cells("A1:G1")

headers_a = ["Priority", "Category", "Description", "Impact", "Suggested Fix", "Owner", "Blocked By"]
for i, h in enumerate(headers_a, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

critical_rows = [
    ("CRITICAL", "Roles/Schema", "Vendedor role not implemented (absent from roles seed + Role enum; no lead/commission entities).",
     "Cannot track sales pipeline or pay commissions; blocks sales team onboarding.",
     "Add VENDEDOR to roles seed + Role enum, add leads & commissions tables, add /api/leads and /api/commissions endpoints with guards.",
     "Backend team", "Client sign-off on commission model"),
    ("CRITICAL", "Roles/Schema", "Capacitador (instructor) role not implemented; no instructor assignment or cohort/session entities.",
     "Cannot assign a live instructor to a course or track cohorts.",
     "Add CAPACITADOR role, cohorts + sessions tables, instructor assignment on courses, /api/cohorts endpoints.",
     "Backend team", "Client confirms cohort model"),
    ("CRITICAL", "Payments/Tax", "No CFDI/SAT invoicing endpoint (invoices table has cfdi_uuid/cfdi_xml_url but no stamp/fetch endpoints).",
     "Cannot issue tax-valid invoices in Mexico; blocks B2B sales.",
     "Integrate Facturama or SW Sapien; add POST /api/invoices/:id/stamp and GET /api/invoices/:id/cfdi endpoints.",
     "Backend team", "Client picks CFDI provider"),
    ("CRITICAL", "Payments", "No payment provider webhook endpoint; checkout creates provider_ref but no callback to mark paid.",
     "Payments stay in pending state forever; certificates never unlock.",
     "Add POST /api/payments/webhook with signature verification for chosen gateway (Conekta recommended).",
     "Backend team", "Client picks payment gateway"),
    ("CRITICAL", "Security", "No refresh token or session revocation flow. JWT is fire-and-forget with no rotation.",
     "Compliance risk for a regulated LMS; can't log out compromised sessions.",
     "Add refresh_tokens table, POST /api/auth/refresh, POST /api/auth/logout with token blacklist.",
     "Backend team", ""),
    ("HIGH", "Quiz", "Quiz max_attempts enforcement and retake cooldown not verified in submit endpoint.",
     "Employee could brute-force pass; violates 3-attempt policy.",
     "Add attempt counter check + cooldown timer in POST /api/quizzes/:id/submit.",
     "Backend team", ""),
    ("HIGH", "Certificates", "Recertification cycle for NOM-specific expiry windows (annual for NOM-030) not modeled.",
     "Certificates may not expire correctly; regulatory non-compliance.",
     "Add validity_months per course; nightly cron to flip expired certs to status=expired.",
     "Backend team", "Client confirms expiry per course"),
    ("HIGH", "Certificates", "No DC-3 folio auto-generation vs manual entry rules documented.",
     "Folios may collide or miss STPS format.",
     "Confirm PC-XXXX-XXXX-XXXX pattern; add sequence generator with STPS registration prefix.",
     "Backend team", "Client STPS registration number"),
    ("HIGH", "Users", "No bulk user import (CSV/XLSX) endpoint. Client Admin onboarding hundreds of employees has no path.",
     "Manual onboarding won't scale past pilot.",
     "Add POST /api/users/bulk-import with async job + error report download.",
     "Backend team", ""),
    ("HIGH", "Analytics", "No role-based analytics scoping for Client (per-company drilldown) documented.",
     "Client dashboards may leak cross-tenant metrics.",
     "Add company_id filter guard on all analytics queries; add per-company drilldown views.",
     "Backend team", ""),
    ("HIGH", "Content", "No SCORM/xAPI ingestion for third-party course content.",
     "Cannot accept industry-standard course packages.",
     "Add SCORM 1.2/2004 player + xAPI LRS endpoint if in scope.",
     "Backend team", "Client confirms need"),
    ("HIGH", "Catalog", "No public course catalog filter by NOM reference or category.",
     "Public/marketing pages can't showcase by NOM.",
     "Add ?nom= and ?category= query params to GET /api/courses/public.",
     "Backend team", ""),
    ("HIGH", "Certificates", "No QR code verification flow documented.",
     "Verifiers can't scan certs; blocks trust model.",
     "Add QR embed in DC-3 PDF pointing to /verify/:folio; document in API.",
     "Backend team", ""),
    ("HIGH", "Certificates", "No STPS DC-3 template compliance test; no digital signature (SAT) on PDFs.",
     "Certs may be rejected by STPS auditors.",
     "Snapshot official DC-3 template; implement layout test; add SAT digital signature step.",
     "Backend team", "Client provides template + signature"),
]
r = 3
for row in critical_rows:
    ws.append(row)
    prio_cell = ws.cell(row=r, column=1)
    if row[0] == "CRITICAL":
        prio_cell.fill = STATUS_FILL["BLOCKED"]
    elif row[0] == "HIGH":
        prio_cell.fill = STATUS_FILL["PARTIAL"]
    prio_cell.font = Font(bold=True)
    r += 1

# blank rows then section B
r += 2
ws.cell(row=r, column=1, value="B. External blockers (waiting on client)")
ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="FFFFFF")
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
headers_b = ["Priority", "Question", "Client Answer", "What we still need", "Owner"]
for i, h in enumerate(headers_b, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
r += 1

blocker_rows = [
    ("HIGH", "Preferred payment gateway", "Open to our recommendation",
     "DECISION: recommend Conekta (cards + OXXO + SPEI, MX-native). Need client sign-off + merchant credentials.", "Ale (client)"),
    ("HIGH", "CFDI / SAT invoicing provider", "Yes, invoicing required",
     "DECISION: recommend Facturama or SW Sapien. Need pick + API keys + RFC of issuing entity.", "Ale (client)"),
    ("HIGH", "DC-3 format details (STPS registration, instructor name/RFC/folio, course codes, signer)", "Certifier agent will confirm",
     "Registration number, instructor details, STPS clave per NOM course, authorized signer name + signature image.", "Certifier agent"),
    ("HIGH", "Course content (slides for 5 NOM courses + quiz banks + validity period)", "Client provides slides; awaiting delivery",
     "ETA on slide decks, quiz question banks with correct answers, validity months per course.", "Ale (client) + certifier"),
    ("HIGH", "Hosting preference (we host vs client hosts)", "Pending",
     "DECISION needed; impacts monthly cost, SLAs, MX data residency setup.", "Ale (client)"),
    ("HIGH", "Pricing model + package prices + seats per tier", "Pending",
     "One-time vs subscription? Prices per package (basic 5 NOM, extras). Seats included per tier.", "Ale (client)"),
]
for row in blocker_rows:
    ws.append(row)
    ws.cell(row=r, column=1).fill = STATUS_FILL["PARTIAL"]
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1

set_widths(ws, [12, 30, 45, 45, 45, 25, 25])
# wrap everything
for row in ws.iter_rows(min_row=2):
    for c in row:
        if c.value is not None:
            c.alignment = WRAP
            if not c.border or c.border.left.style is None:
                c.border = BORDER

wb.save(OUT)
print(f"WROTE {OUT}")
print(f"size: {OUT.stat().st_size} bytes")

# validate reload
from openpyxl import load_workbook
wb2 = load_workbook(OUT)
for s in wb2.sheetnames:
    print(f"  {s}: {wb2[s].max_row} rows x {wb2[s].max_column} cols")
